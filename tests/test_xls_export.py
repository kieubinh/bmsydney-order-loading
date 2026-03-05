from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.types import Order
from app.xls import generate_accrivia_xls


def test_xls_single_sheet_fixed_mapping_no_price_cells() -> None:
    template = load_workbook(Path("data/test_template.xlsx"))
    assert "Sheet1" in template.sheetnames

    order = Order(
        id="order-1",
        queueNumber="CAB-20260304-001",
        submittedAt=datetime(2026, 3, 4, 0, 0, 0),
        status="Submitted",
        customerName="Alice",
        contactNumber="0411000000",
        debtorCode="CASHCAB",
        orderType="Pickup",
        pickupStore="Cabramatta",
        requiredDate=date(2026, 3, 4),
        deliveryAddress="",
        fulfilmentNote="ready at noon",
        lines=[
            {"itemCode": "ITEM0001", "quantity": 3},
            {"itemCode": "ITEM0002", "quantity": 5},
        ],
    )

    workbook = load_workbook(BytesIO(generate_accrivia_xls(order)))

    assert workbook.sheetnames == ["Sheet1"]
    sheet = workbook["Sheet1"]

    assert sheet["A1"].value == "CASHCAB"
    assert sheet["A3"].value == "2026-03-04"
    assert sheet["A12"].value == "ITEM0001"
    assert sheet["B12"].value == 3
    assert sheet["A13"].value == "ITEM0002"
    assert sheet["B13"].value == 5

    assert sheet["D12"].value is None
    assert sheet["D13"].value is None

"""
BMSydney Order Capture → Queue → Accrivia XLS Export System
Flask + SQLite + openpyxl — no external auth deps
"""

import os, json, hashlib, secrets, re, uuid
from datetime import datetime, date, timedelta
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, abort)
import sqlite3, openpyxl
from openpyxl import Workbook
import io, base64

# Use the directory where app.py lives as the base — works on Windows too
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, "templates"),
            static_folder=os.path.join(BASE_DIR, "static"))
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
DB_PATH = os.path.join(BASE_DIR, "bmsydney.db")
XLS_DIR = os.path.join(BASE_DIR, "generated_xls")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(XLS_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ─── Date helpers (storage: YYYY-MM-DD, display: DD/MM/YYYY) ─────────────────

def to_display_date(iso_str):
    if not iso_str:
        return "—"
    try:
        return datetime.strptime(str(iso_str)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return str(iso_str)

def to_iso_date(val):
    if not val:
        return ""
    s = str(val).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    try:
        return datetime.strptime(s, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return s

app.jinja_env.filters["ddate"] = to_display_date

# ─── DB helpers ──────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE,
            phone TEXT UNIQUE,
            password_hash TEXT NOT NULL,
            name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'CUSTOMER',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS item_master (
            item_code TEXT PRIMARY KEY,
            description TEXT NOT NULL,
            active_flag INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS queue_counter (
            queue_key TEXT PRIMARY KEY,
            counter INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_ref TEXT UNIQUE NOT NULL,
            queue_number TEXT,
            customer_id INTEGER REFERENCES users(id),
            order_type TEXT NOT NULL,
            pickup_store TEXT,
            required_date TEXT,
            delivery_address TEXT,
            contact_number TEXT,
            fulfilment_note TEXT,
            job_name TEXT,
            customer_order_no TEXT,
            debtor_code TEXT DEFAULT '1051034',
            source TEXT DEFAULT 'Manual',
            status TEXT DEFAULT 'Draft',
            submitted_at TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS order_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER REFERENCES orders(id),
            line_no INTEGER,
            item_code TEXT NOT NULL,
            description TEXT,
            quantity REAL NOT NULL,
            note TEXT
        );
        CREATE TABLE IF NOT EXISTS order_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER REFERENCES orders(id),
            attach_type TEXT,
            file_path TEXT,
            version_no INTEGER DEFAULT 1,
            created_by INTEGER,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER REFERENCES orders(id),
            actor_id INTEGER REFERENCES users(id),
            action_type TEXT,
            changes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        """)
        seed_data(db)

def seed_data(db):
    # Check if already seeded
    existing = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing > 0:
        return

    def pw(p): return hashlib.sha256(p.encode()).hexdigest()

    # Staff users
    staff = [
        ("cabra_cs@test.com", None, "CS-CABRA", "Cabra CS"),
        ("lid_cs@test.com", None, "CS-LIDCOMBE", "Lidcombe CS"),
        ("ops@test.com", None, "OPS-DELIVERY", "OPS Delivery"),
        ("admin@test.com", None, "ADMIN", "Admin User"),
        ("customer@test.com", "0412345678", "CUSTOMER", "Test Customer"),
    ]
    for email, phone, role, name in staff:
        db.execute("INSERT OR IGNORE INTO users (email,phone,password_hash,name,role) VALUES (?,?,?,?,?)",
                   (email, phone, pw("password123"), name, role))

    # 100 sample items
    items = []
    cats = ["FLATBAR", "ROUNDBAR", "PIPE", "ANGLE", "CHANNEL", "BEAM", "PLATE", "MESH", "TUBE", "SHEET"]
    for i in range(1, 101):
        code = f"ITEM{i:04d}"
        cat = cats[i % len(cats)]
        desc = f"{cat} {i*10}x{i*5} - {i*3}m Length"
        items.append((code, desc))
    # Also add template items
    template_items = [
        ("10080", "Standard Flat Bar 10080"),
        ("100FLATBAR PB", "100mm Flat Bar PB"),
        ("100FLATBAR3.6 CA", "100mm Flat Bar 3.6m CA"),
        ("100SK3.6 CA", "100 SK 3.6m CA"),
    ]
    items.extend(template_items)
    db.executemany("INSERT OR IGNORE INTO item_master (item_code,description) VALUES (?,?)", items)

    # Default settings
    settings = [
        ("CABRA_CS_EMAIL", "cabra_cs@test.com"),
        ("LIDCOMBE_CS_EMAIL", "lid_cs@test.com"),
        ("OPS_DELIVERY_EMAIL", "ops@test.com"),
        ("QUEUE_RESET", "daily"),
        ("DEBTOR_CODE", "1051034"),
    ]
    db.executemany("INSERT OR IGNORE INTO app_settings (key,value) VALUES (?,?)", settings)

# ─── Auth helpers ─────────────────────────────────────────────────────────────

def hash_pw(p): return hashlib.sha256(p.encode()).hexdigest()

def current_user():
    uid = session.get("user_id")
    if not uid: return None
    with get_db() as db:
        return db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            u = current_user()
            if not u or u["role"] not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated
    return decorator

def get_setting(key, default=""):
    with get_db() as db:
        r = db.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

# ─── Queue number ─────────────────────────────────────────────────────────────

def assign_queue_number(order_type, pickup_store):
    today = date.today().strftime("%Y%m%d")
    if order_type == "Pickup":
        prefix = "CAB" if pickup_store == "Cabramatta" else "LID"
    else:
        prefix = "DEL"
    key = f"{prefix}-{today}"
    with get_db() as db:
        db.execute("INSERT OR IGNORE INTO queue_counter (queue_key,counter) VALUES (?,0)", (key,))
        db.execute("UPDATE queue_counter SET counter=counter+1 WHERE queue_key=?", (key,))
        n = db.execute("SELECT counter FROM queue_counter WHERE queue_key=?", (key,)).fetchone()["counter"]
    return f"{prefix}-{today}-{n:03d}"

# ─── XLS generator ────────────────────────────────────────────────────────────

def generate_xls(order, lines):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    today = date.today()
    req_iso = order["required_date"] or today.strftime("%Y-%m-%d")
    req_display = to_display_date(req_iso)

    # Header rows (A=label, B=value)
    ws["A1"] = "Debtor Code";      ws["B1"] = order["debtor_code"] or "1051034"
    ws["A2"] = "Date";             ws["B2"] = today.strftime("%d/%m/%Y")
    ws["A3"] = "Date Required";    ws["B3"] = req_display
    ws["A4"] = "Customer Order No";ws["B4"] = order["customer_order_no"] or ""
    ws["A5"] = "Job Name";         ws["B5"] = order["job_name"] or ""

    if order["order_type"] == "Delivery":
        addr = order["delivery_address"] or ""
        parts = addr.split(",", 2)
        ws["A6"] = "Job Address Line 1"; ws["B6"] = parts[0].strip() if len(parts) > 0 else ""
        ws["A7"] = "Job Address Line 2"; ws["B7"] = parts[1].strip() if len(parts) > 1 else ""
        ws["A8"] = "Job Address Line 3"; ws["B8"] = parts[2].strip() if len(parts) > 2 else ""
    else:
        ws["A6"] = "Job Address Line 1"; ws["B6"] = f"Pickup: {order['pickup_store'] or ''}"
        ws["A7"] = "Job Address Line 2"; ws["B7"] = ""
        ws["A8"] = "Job Address Line 3"; ws["B8"] = ""

    ws["A9"] = "Sales Rep Code"; ws["B9"] = ""

    # Line headers row 11
    ws["A11"] = "Stock Code"
    ws["B11"] = "Description"
    ws["C11"] = "Quan"
    # NOTE: Never write col D (Rate Ex / price)

    # Line items from row 12
    for i, line in enumerate(lines):
        r = 12 + i
        ws.cell(row=r, column=1, value=line["item_code"])
        ws.cell(row=r, column=2, value=line["description"] or "")
        ws.cell(row=r, column=3, value=float(line["quantity"]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── Email (mock/log) ─────────────────────────────────────────────────────────

def send_email_notification(order, lines, xls_bytes, queue_number):
    """In dev, just log. In prod, wire up SMTP."""
    otype = order["order_type"]
    store = order.get("pickup_store")
    if otype == "Pickup" and store == "Cabramatta":
        to = get_setting("CABRA_CS_EMAIL")
    elif otype == "Pickup" and store == "Lidcombe":
        to = get_setting("LIDCOMBE_CS_EMAIL")
    else:
        to = get_setting("OPS_DELIVERY_EMAIL")
    subject = f"[Order Queue #{queue_number}] {otype} - {order['order_ref']}"
    body = f"Order {order['order_ref']} submitted. Queue: {queue_number}. Items: {len(lines)}."
    print(f"[EMAIL] To: {to} | Subject: {subject} | {body}")
    return to

# ─── Routes: Auth ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    if session.get("user_id"):
        u = current_user()
        if u and u["role"] == "CUSTOMER":
            return redirect(url_for("customer_dashboard"))
        elif u:
            return redirect(url_for("staff_queue"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        identifier = request.form.get("identifier","").strip()
        password = request.form.get("password","")
        with get_db() as db:
            u = db.execute(
                "SELECT * FROM users WHERE email=? OR phone=?", (identifier, identifier)
            ).fetchone()
        if u and u["password_hash"] == hash_pw(password):
            session["user_id"] = u["id"]
            session["user_role"] = u["role"]
            session["user_name"] = u["name"]
            if u["role"] == "CUSTOMER":
                return redirect(url_for("customer_dashboard"))
            return redirect(url_for("staff_queue"))
        flash("Invalid credentials. Please try again.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─── Routes: Customer ─────────────────────────────────────────────────────────

@app.route("/customer/dashboard")
@login_required
def customer_dashboard():
    u = current_user()
    with get_db() as db:
        orders = db.execute(
            "SELECT * FROM orders WHERE customer_id=? ORDER BY created_at DESC", (u["id"],)
        ).fetchall()
    return render_template("customer_dashboard.html", user=u, orders=orders)

@app.route("/customer/new-order")
@login_required
def new_order():
    u = current_user()
    with get_db() as db:
        items = db.execute("SELECT * FROM item_master WHERE active_flag=1 ORDER BY item_code").fetchall()
    return render_template("new_order.html", user=u, items=items)

@app.route("/api/items")
@login_required
def api_items():
    q = request.args.get("search","").strip()
    with get_db() as db:
        if q:
            items = db.execute(
                "SELECT * FROM item_master WHERE active_flag=1 AND (item_code LIKE ? OR description LIKE ?) LIMIT 50",
                (f"%{q}%", f"%{q}%")
            ).fetchall()
        else:
            items = db.execute("SELECT * FROM item_master WHERE active_flag=1 ORDER BY item_code LIMIT 100").fetchall()
    return jsonify([dict(i) for i in items])

@app.route("/api/orders", methods=["POST"])
@login_required
def create_order():
    u = current_user()
    data = request.get_json()
    ref = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(3).upper()}"
    with get_db() as db:
        db.execute("""INSERT INTO orders 
            (order_ref, customer_id, order_type, pickup_store, required_date,
             delivery_address, contact_number, fulfilment_note, job_name,
             customer_order_no, source, status, debtor_code)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (ref, u["id"], data.get("order_type"), data.get("pickup_store"),
             to_iso_date(data.get("required_date","")), data.get("delivery_address"),
             data.get("contact_number"), data.get("fulfilment_note"),
             data.get("job_name"), data.get("customer_order_no"),
             data.get("source","Manual"), "Draft",
             get_setting("DEBTOR_CODE","1051034")))
        order_id = db.execute("SELECT id FROM orders WHERE order_ref=?", (ref,)).fetchone()["id"]
        lines = data.get("lines", [])
        for i, line in enumerate(lines):
            db.execute("INSERT INTO order_lines (order_id,line_no,item_code,description,quantity) VALUES (?,?,?,?,?)",
                       (order_id, i+1, line["item_code"], line.get("description",""), float(line["quantity"])))
        db.execute("INSERT INTO audit_log (order_id,actor_id,action_type,changes) VALUES (?,?,?,?)",
                   (order_id, u["id"], "Create", json.dumps({"status":"Draft"})))
    return jsonify({"success": True, "order_id": order_id, "order_ref": ref})

@app.route("/api/orders/<int:order_id>/submit", methods=["POST"])
@login_required
def submit_order(order_id):
    u = current_user()
    with get_db() as db:
        order = db.execute("SELECT * FROM orders WHERE id=? AND customer_id=?", (order_id, u["id"])).fetchone()
        if not order:
            return jsonify({"error": "Order not found"}), 404
        if order["status"] != "Draft":
            return jsonify({"error": "Order already submitted"}), 400
        lines = db.execute("SELECT * FROM order_lines WHERE order_id=? ORDER BY line_no", (order_id,)).fetchall()
        if not lines:
            return jsonify({"error": "No items in order"}), 400

        queue_num = assign_queue_number(order["order_type"], order["pickup_store"])
        now = datetime.now().isoformat()
        db.execute("UPDATE orders SET queue_number=?, status='Submitted', submitted_at=?, updated_at=? WHERE id=?",
                   (queue_num, now, now, order_id))
        xls_bytes = generate_xls(dict(order), [dict(l) for l in lines])
        fname = f"{order['order_ref']}_v1.xlsx"
        fpath = os.path.join(XLS_DIR, fname)
        with open(fpath, "wb") as f:
            f.write(xls_bytes)
        db.execute("INSERT INTO order_attachments (order_id,attach_type,file_path,version_no,created_by) VALUES (?,?,?,?,?)",
                   (order_id, "GeneratedXLS", fname, 1, u["id"]))
        db.execute("INSERT INTO audit_log (order_id,actor_id,action_type,changes) VALUES (?,?,?,?)",
                   (order_id, u["id"], "Submit", json.dumps({"queue_number": queue_num})))
        order_dict = dict(db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone())

    send_email_notification(order_dict, [dict(l) for l in lines], xls_bytes, queue_num)
    return jsonify({"success": True, "queue_number": queue_num, "order_ref": order_dict["order_ref"]})

@app.route("/customer/order/<int:order_id>")
@login_required
def order_detail_customer(order_id):
    u = current_user()
    with get_db() as db:
        order = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order or (u["role"] == "CUSTOMER" and order["customer_id"] != u["id"]):
            abort(404)
        lines = db.execute("SELECT * FROM order_lines WHERE order_id=? ORDER BY line_no", (order_id,)).fetchall()
        attachments = db.execute("SELECT * FROM order_attachments WHERE order_id=? ORDER BY version_no", (order_id,)).fetchall()
        customer = db.execute("SELECT * FROM users WHERE id=?", (order["customer_id"],)).fetchone()
    return render_template("order_detail.html", order=order, lines=lines, attachments=attachments,
                           customer=customer, user=u, is_staff=u["role"] != "CUSTOMER")

# ─── Routes: Staff ────────────────────────────────────────────────────────────

@app.route("/staff/queue")
@login_required
def staff_queue():
    u = current_user()
    if u["role"] == "CUSTOMER":
        return redirect(url_for("customer_dashboard"))
    with get_db() as db:
        if u["role"] == "CS-CABRA":
            orders = db.execute(
                "SELECT o.*, u.name as customer_name FROM orders o JOIN users u ON o.customer_id=u.id "
                "WHERE o.order_type='Pickup' AND o.pickup_store='Cabramatta' AND o.status NOT IN ('Completed','Cancelled') "
                "ORDER BY o.queue_number"
            ).fetchall()
        elif u["role"] == "CS-LIDCOMBE":
            orders = db.execute(
                "SELECT o.*, u.name as customer_name FROM orders o JOIN users u ON o.customer_id=u.id "
                "WHERE o.order_type='Pickup' AND o.pickup_store='Lidcombe' AND o.status NOT IN ('Completed','Cancelled') "
                "ORDER BY o.queue_number"
            ).fetchall()
        elif u["role"] == "OPS-DELIVERY":
            orders = db.execute(
                "SELECT o.*, u.name as customer_name FROM orders o JOIN users u ON o.customer_id=u.id "
                "WHERE o.order_type='Delivery' AND o.status NOT IN ('Completed','Cancelled') "
                "ORDER BY o.queue_number"
            ).fetchall()
        else:  # ADMIN
            orders = db.execute(
                "SELECT o.*, u.name as customer_name FROM orders o JOIN users u ON o.customer_id=u.id "
                "WHERE o.status NOT IN ('Completed','Cancelled') ORDER BY o.queue_number"
            ).fetchall()
    return render_template("staff_queue.html", orders=orders, user=u)

@app.route("/staff/order/<int:order_id>", methods=["GET","POST"])
@login_required
def staff_order_detail(order_id):
    u = current_user()
    if u["role"] == "CUSTOMER":
        abort(403)
    with get_db() as db:
        order = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order: abort(404)
        if request.method == "POST":
            data = request.form
            old = dict(order)
            db.execute("""UPDATE orders SET
                required_date=?, delivery_address=?, contact_number=?,
                fulfilment_note=?, job_name=?, customer_order_no=?,
                updated_at=datetime('now') WHERE id=?""",
                (to_iso_date(data.get("required_date","")), data.get("delivery_address"),
                 data.get("contact_number"), data.get("fulfilment_note"),
                 data.get("job_name"), data.get("customer_order_no"), order_id))
            changes = {k: {"old": old.get(k), "new": data.get(k)}
                       for k in ["required_date","delivery_address","contact_number","fulfilment_note"]
                       if old.get(k) != data.get(k)}
            db.execute("INSERT INTO audit_log (order_id,actor_id,action_type,changes) VALUES (?,?,?,?)",
                       (order_id, u["id"], "Update", json.dumps(changes)))
            flash("Order updated successfully.", "success")
            return redirect(url_for("staff_order_detail", order_id=order_id))

        lines = db.execute("SELECT * FROM order_lines WHERE order_id=? ORDER BY line_no", (order_id,)).fetchall()
        attachments = db.execute("SELECT * FROM order_attachments WHERE order_id=? ORDER BY version_no DESC", (order_id,)).fetchall()
        audit = db.execute(
            "SELECT a.*, u.name as actor_name FROM audit_log a LEFT JOIN users u ON a.actor_id=u.id "
            "WHERE a.order_id=? ORDER BY a.created_at DESC", (order_id,)
        ).fetchall()
        customer = db.execute("SELECT * FROM users WHERE id=?", (order["customer_id"],)).fetchone()
        items = db.execute("SELECT * FROM item_master WHERE active_flag=1 ORDER BY item_code").fetchall()
    return render_template("staff_order_detail.html", order=order, lines=lines,
                           attachments=attachments, audit=audit, customer=customer,
                           user=u, items=items)

@app.route("/staff/order/<int:order_id>/status", methods=["POST"])
@login_required
def update_status(order_id):
    u = current_user()
    if u["role"] == "CUSTOMER": abort(403)
    new_status = request.form.get("status")
    allowed = ["InReview","Confirmed","Exported","Completed","Cancelled"]
    if new_status not in allowed:
        flash("Invalid status.", "error")
        return redirect(url_for("staff_order_detail", order_id=order_id))
    with get_db() as db:
        old = db.execute("SELECT status FROM orders WHERE id=?", (order_id,)).fetchone()["status"]
        db.execute("UPDATE orders SET status=?, updated_at=datetime('now') WHERE id=?", (new_status, order_id))
        db.execute("INSERT INTO audit_log (order_id,actor_id,action_type,changes) VALUES (?,?,?,?)",
                   (order_id, u["id"], "StatusChange", json.dumps({"old": old, "new": new_status})))
    flash(f"Status updated to {new_status}.", "success")
    return redirect(url_for("staff_order_detail", order_id=order_id))

@app.route("/staff/order/<int:order_id>/update-lines", methods=["POST"])
@login_required
def update_lines(order_id):
    u = current_user()
    if u["role"] == "CUSTOMER": abort(403)
    data = request.get_json()
    lines = data.get("lines", [])
    with get_db() as db:
        db.execute("DELETE FROM order_lines WHERE order_id=?", (order_id,))
        for i, line in enumerate(lines):
            db.execute("INSERT INTO order_lines (order_id,line_no,item_code,description,quantity) VALUES (?,?,?,?,?)",
                       (order_id, i+1, line["item_code"], line.get("description",""), float(line["quantity"])))
        db.execute("INSERT INTO audit_log (order_id,actor_id,action_type,changes) VALUES (?,?,?,?)",
                   (order_id, u["id"], "UpdateLines", json.dumps({"lines_count": len(lines)})))
    return jsonify({"success": True})

@app.route("/staff/order/<int:order_id>/generate-xls", methods=["POST"])
@login_required
def generate_xls_route(order_id):
    u = current_user()
    if u["role"] == "CUSTOMER": abort(403)
    with get_db() as db:
        order = db.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        lines = db.execute("SELECT * FROM order_lines WHERE order_id=? ORDER BY line_no", (order_id,)).fetchall()
        max_v = db.execute("SELECT MAX(version_no) as mv FROM order_attachments WHERE order_id=? AND attach_type='GeneratedXLS'", (order_id,)).fetchone()["mv"] or 0
        new_v = max_v + 1
        xls_bytes = generate_xls(dict(order), [dict(l) for l in lines])
        fname = f"{order['order_ref']}_v{new_v}.xlsx"
        fpath = os.path.join(XLS_DIR, fname)
        with open(fpath, "wb") as f:
            f.write(xls_bytes)
        db.execute("INSERT INTO order_attachments (order_id,attach_type,file_path,version_no,created_by) VALUES (?,?,?,?,?)",
                   (order_id, "GeneratedXLS", fname, new_v, u["id"]))
        db.execute("INSERT INTO audit_log (order_id,actor_id,action_type,changes) VALUES (?,?,?,?)",
                   (order_id, u["id"], "GenerateXLS", json.dumps({"version": new_v})))
    flash(f"XLS v{new_v} generated successfully.", "success")
    return redirect(url_for("staff_order_detail", order_id=order_id))

@app.route("/download/xls/<int:attachment_id>")
@login_required
def download_xls(attachment_id):
    with get_db() as db:
        att = db.execute("SELECT * FROM order_attachments WHERE id=?", (attachment_id,)).fetchone()
        if not att: abort(404)
        fpath = os.path.join(XLS_DIR, att["file_path"])
        if not os.path.exists(fpath): abort(404)
    return send_file(fpath, as_attachment=True, download_name=att["file_path"])

@app.route("/api/orders/<int:order_id>/upload-scan", methods=["POST"])
@login_required
def upload_scan(order_id):
    u = current_user()
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    fname = f"{order_id}_{secrets.token_hex(4)}_{f.filename}"
    fpath = os.path.join(UPLOAD_DIR, fname)
    f.save(fpath)
    with get_db() as db:
        db.execute("INSERT INTO order_attachments (order_id,attach_type,file_path,version_no,created_by) VALUES (?,?,?,?,?)",
                   (order_id, "ScanImage", fname, 1, u["id"]))
    # MVP: return empty extracted lines (stub — real OCR would go here)
    return jsonify({"success": True, "extracted_lines": [], "message": "Image uploaded. Please enter items manually."})

@app.route("/admin/settings", methods=["GET","POST"])
@login_required
def admin_settings():
    u = current_user()
    if u["role"] != "ADMIN": abort(403)
    with get_db() as db:
        if request.method == "POST":
            for key in ["CABRA_CS_EMAIL","LIDCOMBE_CS_EMAIL","OPS_DELIVERY_EMAIL","DEBTOR_CODE","QUEUE_RESET"]:
                val = request.form.get(key, "")
                db.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES (?,?)", (key, val))
            flash("Settings saved.", "success")
        settings = {r["key"]: r["value"] for r in db.execute("SELECT * FROM app_settings").fetchall()}
        users = db.execute("SELECT * FROM users ORDER BY role,name").fetchall()
        items = db.execute("SELECT * FROM item_master ORDER BY item_code").fetchall()
    return render_template("admin_settings.html", settings=settings, user=u, users=users, items=items)

@app.route("/admin/items/toggle/<item_code>", methods=["POST"])
@login_required
def toggle_item(item_code):
    u = current_user()
    if u["role"] != "ADMIN": abort(403)
    with get_db() as db:
        db.execute("UPDATE item_master SET active_flag = 1 - active_flag WHERE item_code=?", (item_code,))
    return redirect(url_for("admin_settings"))

if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)
